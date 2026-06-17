from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func
from typing import Optional, List, Tuple
from datetime import datetime, date, timedelta
from io import BytesIO

from app.database import get_db
from app.services import ReportService
from app.models import (
    DailyReport,
    ReportRecord,
    Donor,
    Organ,
    Recipient,
    Allocation,
    Surgery,
    Transport,
    TransportAlert,
    FollowUpRecord,
    FollowUpAlert,
    Consumable,
    Approval,
    WaitingList,
)
from app.schemas import (
    ReportType,
    ExportFormat,
    DailyReportStats,
    DailyReportResponse,
    ReportExportRequest,
    PaginatedResponse,
)

router = APIRouter(prefix="/api/reports", tags=["reports"])


class ReportRouterService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.service = ReportService(db) if ReportService else None

    async def generate_daily_report(
        self,
        report_date: date,
        province: Optional[str] = None,
        transplant_center_id: Optional[str] = None,
    ) -> DailyReport:
        start_of_day = datetime.combine(report_date, datetime.min.time())
        end_of_day = datetime.combine(report_date, datetime.max.time())

        stats = DailyReportStats()

        donor_query = select(Donor).where(
            Donor.created_at >= start_of_day,
            Donor.created_at <= end_of_day,
        )
        if province:
            donor_query = donor_query.where(Donor.province == province)
        new_donors = list((await self.db.execute(donor_query)).scalars().all())
        stats.new_donor_count = len(new_donors)
        stats.donor_approved_count = len([d for d in new_donors if d.status == "verified"])

        organ_query = select(Organ).where(
            Organ.created_at >= start_of_day,
            Organ.created_at <= end_of_day,
        )
        new_organs = list((await self.db.execute(organ_query)).scalars().all())
        stats.new_organ_count = len(new_organs)

        organ_available_query = select(Organ).where(Organ.status == "available")
        stats.organ_available_count = (await self.db.execute(
            select(func.count()).select_from(organ_available_query.subquery())
        )).scalar() or 0

        recipient_query = select(Recipient).where(
            Recipient.created_at >= start_of_day,
            Recipient.created_at <= end_of_day,
        )
        if transplant_center_id:
            recipient_query = recipient_query.where(Recipient.transplant_center_id == transplant_center_id)
        new_recipients = list((await self.db.execute(recipient_query)).scalars().all())
        stats.new_recipient_count = len(new_recipients)

        waiting_query = select(WaitingList)
        stats.waiting_list_count = (await self.db.execute(
            select(func.count()).select_from(waiting_query.subquery())
        )).scalar() or 0

        allocation_query = select(Allocation).where(
            Allocation.created_at >= start_of_day,
            Allocation.created_at <= end_of_day,
        )
        new_allocations = list((await self.db.execute(allocation_query)).scalars().all())
        stats.allocation_request_count = len(new_allocations)
        stats.allocation_approved_count = len(
            [a for a in new_allocations if a.status in ("provincial_approved", "national_approved")]
        )

        surgery_query = select(Surgery).where(
            Surgery.surgery_date >= start_of_day,
            Surgery.surgery_date <= end_of_day,
            Surgery.surgery_status == "completed",
        )
        if transplant_center_id:
            surgery_query = surgery_query.where(Surgery.transplant_center_id == transplant_center_id)
        stats.transplant_count = (await self.db.execute(
            select(func.count()).select_from(surgery_query.subquery())
        )).scalar() or 0

        transport_query = select(Transport).where(
            Transport.created_at >= start_of_day,
            Transport.created_at <= end_of_day,
        )
        new_transports = list((await self.db.execute(transport_query)).scalars().all())
        stats.transport_count = len(new_transports)
        stats.transport_on_time_count = len(
            [t for t in new_transports if t.status == "delivered"]
        )

        transport_alert_query = select(TransportAlert).where(
            TransportAlert.alert_time >= start_of_day,
            TransportAlert.alert_time <= end_of_day,
        )
        stats.transport_alert_count = (await self.db.execute(
            select(func.count()).select_from(transport_alert_query.subquery())
        )).scalar() or 0

        followup_query = select(FollowUpRecord).where(
            FollowUpRecord.followup_date >= start_of_day,
            FollowUpRecord.followup_date <= end_of_day,
        )
        stats.followup_count = (await self.db.execute(
            select(func.count()).select_from(followup_query.subquery())
        )).scalar() or 0

        followup_alert_query = select(FollowUpAlert).where(
            FollowUpAlert.alert_time >= start_of_day,
            FollowUpAlert.alert_time <= end_of_day,
        )
        stats.followup_alert_count = (await self.db.execute(
            select(func.count()).select_from(followup_alert_query.subquery())
        )).scalar() or 0

        low_stock_query = select(Consumable).where(Consumable.status.in_(["low", "critical"]))
        stats.consumable_low_stock_count = (await self.db.execute(
            select(func.count()).select_from(low_stock_query.subquery())
        )).scalar() or 0

        approval_pending_query = select(Approval).where(Approval.status == "pending")
        stats.approval_pending_count = (await self.db.execute(
            select(func.count()).select_from(approval_pending_query.subquery())
        )).scalar() or 0

        organ_distribution: dict[str, int] = {}
        for organ in new_organs:
            ot = str(organ.organ_type)
            organ_distribution[ot] = organ_distribution.get(ot, 0) + 1

        urgency_distribution: dict[str, int] = {}
        waiting_with_urgency = list((await self.db.execute(
            select(WaitingList)
        )).scalars().all())
        for wl in waiting_with_urgency:
            ul = str(wl.urgency_level)
            urgency_distribution[ul] = urgency_distribution.get(ul, 0) + 1

        existing = await self.db.execute(
            select(DailyReport).where(
                DailyReport.report_date == report_date,
                DailyReport.province == province,
                DailyReport.transplant_center_id == transplant_center_id,
            )
        )
        report = existing.scalar_one_or_none()
        if report:
            report.generated_at = datetime.utcnow()
            report.stats = stats.model_dump()
            report.organ_distribution = organ_distribution
            report.urgency_distribution = urgency_distribution
        else:
            report = DailyReport(
                report_date=report_date,
                report_type="daily",
                generated_at=datetime.utcnow(),
                province=province,
                transplant_center_id=transplant_center_id,
                stats=stats.model_dump(),
                organ_distribution=organ_distribution,
                urgency_distribution=urgency_distribution,
            )
            self.db.add(report)

        await self.db.flush()
        await self.db.refresh(report)
        return report

    async def get_report(self, report_id: str) -> Optional[DailyReport]:
        result = await self.db.execute(
            select(DailyReport).where(DailyReport.id == report_id)
        )
        return result.scalar_one_or_none()

    async def list_reports(
        self,
        skip: int = 0,
        limit: int = 100,
        report_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        province: Optional[str] = None,
        transplant_center_id: Optional[str] = None,
    ) -> Tuple[List[DailyReport], int]:
        query = select(DailyReport)
        if report_type:
            query = query.where(DailyReport.report_type == report_type)
        if start_date:
            query = query.where(DailyReport.report_date >= start_date)
        if end_date:
            query = query.where(DailyReport.report_date <= end_date)
        if province:
            query = query.where(DailyReport.province == province)
        if transplant_center_id:
            query = query.where(DailyReport.transplant_center_id == transplant_center_id)

        count_query = select(func.count()).select_from(query.subquery())
        total = (await self.db.execute(count_query)).scalar() or 0

        query = query.order_by(DailyReport.report_date.desc()).offset(skip).limit(limit)
        result = await self.db.execute(query)
        items = list(result.scalars().all())

        return items, total

    async def query_by_date_range(
        self,
        start_date: date,
        end_date: date,
        province: Optional[str] = None,
        transplant_center_id: Optional[str] = None,
    ) -> list[dict]:
        if end_date < start_date:
            raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")

        delta = end_date - start_date
        if delta.days > 365:
            raise HTTPException(status_code=400, detail="查询范围不能超过365天")

        reports = []
        current_date = start_date
        while current_date <= end_date:
            report = await self.generate_daily_report(
                current_date, province, transplant_center_id
            )
            reports.append({
                "report_date": report.report_date.isoformat(),
                "stats": report.stats,
                "organ_distribution": report.organ_distribution,
                "urgency_distribution": report.urgency_distribution,
            })
            current_date += timedelta(days=1)

        return reports

    async def export_to_excel(
        self,
        data: ReportExportRequest,
    ) -> BytesIO:
        import json

        reports_data = await self.query_by_date_range(
            start_date=data.start_date,
            end_date=data.end_date,
            province=data.province,
            transplant_center_id=data.transplant_center_id,
        )

        output = BytesIO()
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()

            ws = wb.active
            ws.title = "报表汇总"

            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=12, color="FFFFFF")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            headers = [
                "日期", "新增捐赠者", "通过审核捐赠者", "新增器官",
                "可用器官数", "新增受赠者", "等待名单数",
                "分配申请数", "分配通过数", "移植手术数",
                "运输任务数", "准时运输数", "运输告警数",
                "随访记录数", "随访告警数", "低库存耗材数", "待审批数",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row_idx, report in enumerate(reports_data, 2):
                stats = report.get("stats", {})
                values = [
                    report.get("report_date", ""),
                    stats.get("new_donor_count", 0),
                    stats.get("donor_approved_count", 0),
                    stats.get("new_organ_count", 0),
                    stats.get("organ_available_count", 0),
                    stats.get("new_recipient_count", 0),
                    stats.get("waiting_list_count", 0),
                    stats.get("allocation_request_count", 0),
                    stats.get("allocation_approved_count", 0),
                    stats.get("transplant_count", 0),
                    stats.get("transport_count", 0),
                    stats.get("transport_on_time_count", 0),
                    stats.get("transport_alert_count", 0),
                    stats.get("followup_count", 0),
                    stats.get("followup_alert_count", 0),
                    stats.get("consumable_low_stock_count", 0),
                    stats.get("approval_pending_count", 0),
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 15

            ws2 = wb.create_sheet("器官分布")
            organ_headers = ["日期"]
            all_organ_types = set()
            for report in reports_data:
                od = report.get("organ_distribution", {}) or {}
                all_organ_types.update(od.keys())
            organ_headers.extend(sorted(all_organ_types))

            for col, header in enumerate(organ_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row_idx, report in enumerate(reports_data, 2):
                ws2.cell(row=row_idx, column=1, value=report.get("report_date", ""))
                od = report.get("organ_distribution", {}) or {}
                for col_idx, ot in enumerate(sorted(all_organ_types), 2):
                    ws2.cell(row=row_idx, column=col_idx, value=od.get(ot, 0))

            ws3 = wb.create_sheet("紧急程度分布")
            urgency_headers = ["日期"]
            all_urgency = set()
            for report in reports_data:
                ud = report.get("urgency_distribution", {}) or {}
                all_urgency.update(ud.keys())
            urgency_headers.extend(sorted(all_urgency))

            for col, header in enumerate(urgency_headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row_idx, report in enumerate(reports_data, 2):
                ws3.cell(row=row_idx, column=1, value=report.get("report_date", ""))
                ud = report.get("urgency_distribution", {}) or {}
                for col_idx, ul in enumerate(sorted(all_urgency), 2):
                    ws3.cell(row=row_idx, column=col_idx, value=ud.get(ul, 0))

            wb.save(output)
            output.seek(0)

        except ImportError:
            import csv
            content = "日期,新增捐赠者,通过审核,新增器官,可用器官,新增受赠者,等待名单,分配申请,分配通过,移植数\n"
            for report in reports_data:
                stats = report.get("stats", {})
                content += f"{report.get('report_date','')},{stats.get('new_donor_count',0)},{stats.get('donor_approved_count',0)},{stats.get('new_organ_count',0)},{stats.get('organ_available_count',0)},{stats.get('new_recipient_count',0)},{stats.get('waiting_list_count',0)},{stats.get('allocation_request_count',0)},{stats.get('allocation_approved_count',0)},{stats.get('transplant_count',0)}\n"
            output.write(content.encode("utf-8-sig"))
            output.seek(0)

        return output


@router.post("/daily/generate")
async def generate_daily_report(
    report_date: date = Query(..., description="报表日期"),
    province: Optional[str] = None,
    transplant_center_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    service = ReportRouterService(db)
    report = await service.generate_daily_report(report_date, province, transplant_center_id)
    return {"code": 200, "message": "日报生成成功", "data": report}


@router.get("/{report_id}")
async def get_report(
    report_id: str,
    db: AsyncSession = Depends(get_db),
):
    service = ReportRouterService(db)
    report = await service.get_report(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="报表不存在")
    return {"code": 200, "message": "success", "data": report}


@router.get("")
async def list_reports(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    report_type: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    province: Optional[str] = None,
    transplant_center_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    service = ReportRouterService(db)
    skip = (page - 1) * size
    items, total = await service.list_reports(
        skip=skip,
        limit=size,
        report_type=report_type,
        start_date=start_date,
        end_date=end_date,
        province=province,
        transplant_center_id=transplant_center_id,
    )
    total_pages = (total + size - 1) // size
    return {
        "code": 200,
        "message": "success",
        "data": {
            "items": items,
            "total": total,
            "page": page,
            "page_size": size,
            "total_pages": total_pages,
        },
    }


@router.get("/query/range")
async def query_by_date_range(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    province: Optional[str] = None,
    transplant_center_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    service = ReportRouterService(db)
    data = await service.query_by_date_range(start_date, end_date, province, transplant_center_id)
    return {"code": 200, "message": "success", "data": data}


@router.post("/export/excel")
async def export_to_excel(
    data: ReportExportRequest,
    db: AsyncSession = Depends(get_db),
):
    service = ReportRouterService(db)
    file_data = await service.export_to_excel(data)

    filename = f"report_{data.start_date}_{data.end_date}.xlsx"
    if data.export_format == "csv":
        filename = f"report_{data.start_date}_{data.end_date}.csv"
        media_type = "text/csv"
    else:
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([file_data.getvalue()]),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
