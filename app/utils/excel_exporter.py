import os
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelExport:
    PROVINCE_HEADERS = [
        "省份",
        "捐献量",
        "分配成功率",
        "运输准时率",
        "受者存活率",
        "耗材消耗",
    ]

    CENTER_HEADERS = [
        "中心名称",
        "省份",
        "捐献量",
        "分配成功率",
        "运输准时率",
        "受者存活率",
        "耗材消耗",
    ]

    def __init__(self) -> None:
        self._header_font = Font(bold=True, size=11)
        self._header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self._header_font_white = Font(bold=True, size=11, color="FFFFFF")
        self._center_align = Alignment(horizontal="center", vertical="center")
        self._thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def _write_headers(self, ws, headers: List[str]) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self._header_font_white
            cell.fill = self._header_fill
            cell.alignment = self._center_align
            cell.border = self._thin_border

    def _write_rows(self, ws, rows: List[List[Any]], start_row: int = 2) -> None:
        for row_idx, row_data in enumerate(rows, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = self._center_align
                cell.border = self._thin_border

    def _auto_column_width(self, ws) -> None:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_daily_report(self, report_data: Dict, filepath: str) -> str:
        wb = Workbook()

        province_ws = wb.active
        province_ws.title = "省份统计"
        self._write_headers(province_ws, self.PROVINCE_HEADERS)

        province_data: List[List[Any]] = []
        province_stats = report_data.get("province_stats", [])
        if isinstance(province_stats, dict):
            for province, stats in province_stats.items():
                if isinstance(stats, dict):
                    province_data.append([
                        province,
                        stats.get("donation_count", 0),
                        stats.get("allocation_success_rate", 0),
                        stats.get("transport_on_time_rate", 0),
                        stats.get("recipient_survival_rate", 0),
                        stats.get("consumable_usage", 0),
                    ])
                else:
                    province_data.append([province, stats, 0, 0, 0, 0])
        elif isinstance(province_stats, list):
            for item in province_stats:
                if isinstance(item, dict):
                    province_data.append([
                        item.get("province", ""),
                        item.get("donation_count", 0),
                        item.get("allocation_success_rate", 0),
                        item.get("transport_on_time_rate", 0),
                        item.get("recipient_survival_rate", 0),
                        item.get("consumable_usage", 0),
                    ])

        self._write_rows(province_ws, province_data)
        self._auto_column_width(province_ws)

        center_ws = wb.create_sheet(title="中心统计")
        self._write_headers(center_ws, self.CENTER_HEADERS)

        center_data: List[List[Any]] = []
        center_stats = report_data.get("center_stats", [])
        if isinstance(center_stats, dict):
            for center_id, stats in center_stats.items():
                if isinstance(stats, dict):
                    center_data.append([
                        stats.get("center_name", center_id),
                        stats.get("province", ""),
                        stats.get("donation_count", 0),
                        stats.get("allocation_success_rate", 0),
                        stats.get("transport_on_time_rate", 0),
                        stats.get("recipient_survival_rate", 0),
                        stats.get("consumable_usage", 0),
                    ])
                else:
                    center_data.append([center_id, "", stats, 0, 0, 0, 0])
        elif isinstance(center_stats, list):
            for item in center_stats:
                if isinstance(item, dict):
                    center_data.append([
                        item.get("center_name", item.get("center_id", "")),
                        item.get("province", ""),
                        item.get("donation_count", 0),
                        item.get("allocation_success_rate", 0),
                        item.get("transport_on_time_rate", 0),
                        item.get("recipient_survival_rate", 0),
                        item.get("consumable_usage", 0),
                    ])

        self._write_rows(center_ws, center_data)
        self._auto_column_width(center_ws)

        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        wb.save(filepath)
        return filepath
